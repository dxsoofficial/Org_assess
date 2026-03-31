import os
import subprocess
import argparse
import re
import platform
import ipaddress
import time
from pathlib import Path
import xml.etree.ElementTree as ET

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
except ImportError:
    print("[!] openpyxl is not installed. Run 'pip install openpyxl' for Excel Matrix generation.")
    Workbook = None

def extract_ips(text):
    """Extract valid IP addresses from text using regex."""
    ip_pattern = re.compile(r'(?:[0-9]{1,3}\.){3}[0-9]{1,3}')
    return set(ip_pattern.findall(text))

def run_command(cmd, timeout=300):
    try:
        print(f"[*] Executing: {cmd}")
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=timeout)
        return result.stdout
    except subprocess.TimeoutExpired:
        print(f"[!] Warning: Command timed out: {cmd}")
        return ""
    except subprocess.CalledProcessError as e:
        print(f"[!] Warning: Command failed or returned non-zero exit status: {cmd}")
        return e.stdout if e.stdout else ""
    except Exception as e:
        print(f"[!] Error: {e}")
        return ""

def discover_local_subnets():
    """Automatically discover connected subnets based on local interfaces."""
    print("\n=== PHASE 0: Auto-Discovering Local Subnets ===")
    subnets = set()
    
    # 1. Try ip addr (Linux)
    try:
        out = subprocess.check_output("ip -o -f inet addr show", shell=True, text=True, stderr=subprocess.DEVNULL)
        for line in out.splitlines():
            match = re.search(r'inet\s+([0-9]+\.[0-9]+\.[0-9]+\.[0-9]+/[0-9]+)', line)
            if match:
                ip_cidr = match.group(1)
                if not ip_cidr.startswith("127."):
                    network = ipaddress.ip_network(ip_cidr, strict=False)
                    subnets.add(str(network))
    except Exception:
        pass
        
    # 2. Try ifconfig (macOS / older Linux)
    if not subnets:
        try:
            out = subprocess.check_output("ifconfig", shell=True, text=True, stderr=subprocess.DEVNULL)
            lines = out.splitlines()
            for line in lines:
                if "inet " in line and "127.0.0.1" not in line:
                    parts = line.split()
                    try:
                        idx_inet = parts.index("inet")
                        ip_val = parts[idx_inet + 1]
                        
                        if "netmask" in parts:
                            idx_nm = parts.index("netmask")
                            nm_val = parts[idx_nm + 1]
                            
                            if nm_val.startswith("0x"): # Hex mask (macOS)
                                nm_int = int(nm_val, 16)
                                mask = bin(nm_int).count('1')
                            else: # Dotted decimal
                                mask = sum(bin(int(x)).count('1') for x in nm_val.split('.'))
                                
                            network = ipaddress.ip_network(f"{ip_val}/{mask}", strict=False)
                            subnets.add(str(network))
                    except ValueError:
                        continue
        except Exception:
            pass

    # 3. Try Windows route print (fallback)
    if not subnets and platform.system() == "Windows":
        try:
            out = subprocess.check_output("route print", shell=True, text=True, stderr=subprocess.DEVNULL)
            for line in out.splitlines():
                if "On-link" in line:
                    parts = line.split()
                    if len(parts) >= 3:
                        net, mask = parts[0], parts[1]
                        if not net.startswith("224.") and not net.startswith("127.") and net != "255.255.255.255":
                            try:
                                network = ipaddress.ip_network(f"{net}/{mask}", strict=False)
                                subnets.add(str(network))
                            except ValueError:
                                pass
        except Exception:
            pass
            
    return list(subnets)

def generate_port_matrix(live_hosts_file, output_dir):
    if not Workbook: return
    
    print("\n[*] Generating Port Assessment Matrix for discovered hosts...")
    xml_out = os.path.join(output_dir, "nmap_matrix_scan.xml")
    
    # 45 highly specific MSME-focused ports defined by the user
    port_str = "20,21,22,23,25,53,68,69,80,110,123,135,137,138,139,143,161,162,389,443,445,636,1433,2049,2082,2083,2181,2375,3306,3389,4444,5355,5432,5555,5900,5985,5986,6667,7001,7002,8080,8443,9200,11211,27017"
    
    # Run targeted port scan on live hosts
    cmd_nmap = f"nmap -p {port_str} -iL {live_hosts_file} -sS -oX {xml_out}"
    try:
        subprocess.run(cmd_nmap, shell=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception as e:
        print(f"[!] Matrix scan failed: {e}")
        return

    # Parse XML
    try:
        tree = ET.parse(xml_out)
        root = tree.getroot()
    except Exception:
        print("[!] Failed to parse Nmap XML")
        return

    hosts_data = {}
    all_ports = set()

    for host in root.findall('host'):
        ip = "Unknown"
        for addr in host.findall('address'):
            if addr.get('addrtype') == 'ipv4':
                ip = addr.get('addr')
                break
        
        ports_dict = {}
        ports_elem = host.find('ports')
        if ports_elem is not None:
            for port in ports_elem.findall('port'):
                portid_attr = port.get('portid')
                if not portid_attr: continue
                portid = int(portid_attr)
                
                state_elem = port.find('state')
                state = state_elem.get('state') if state_elem is not None else "unknown"
                
                ports_dict[portid] = state
                all_ports.add(portid)
        
        hosts_data[ip] = ports_dict

    if not hosts_data:
        print("[!] No port data found to generate matrix. Subnet may be down or heavily filtered.")
        return

    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Port Matrix"

    # Headers
    sorted_ports = sorted(list(all_ports))
    headers = ["IP Address"] + [f"Port {p}" for p in sorted_ports]
    ws.append(headers)

    # Styles
    red_fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid") # High Risk / Admin / Vuln
    yellow_fill = PatternFill(start_color="FFCC66", end_color="FFCC66", fill_type="solid") # Medium Risk / Context
    green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid") # Closed or Safe Open

    # Risk Dictionaries (Remaining ports assumed red/vulnerable by default if open)
    med_ports = {53, 80, 443, 8080, 8443, 1433, 3306, 5432, 27017}
    safe_open_ports = {68, 123, 5355}

    for idx, (ip, pdata) in enumerate(hosts_data.items(), start=2):
        row = [ip]
        for port in sorted_ports:
            state = pdata.get(port, "closed")
            row.append(state.upper())
        ws.append(row)

        # Apply formatting
        for col_idx, port in enumerate(sorted_ports, start=2):
            cell = ws.cell(row=idx, column=col_idx)
            val = str(cell.value).upper()
            if "OPEN" in val:
                if port in safe_open_ports:
                    cell.fill = green_fill
                elif port in med_ports:
                    cell.fill = yellow_fill
                else:
                    cell.fill = red_fill # Anything else in this dangerous list that is OPEN is Red
            else:
                cell.fill = green_fill # Closed ports are safe (Green)

    excel_out = os.path.join(output_dir, "port_matrix_assessment.xlsx")
    try:
        wb.save(excel_out)
        print(f"[+] Matrix Assessment generated successfully: {excel_out}")
    except Exception as e:
        print(f"[!] Failed to save Excel Matrix: {e}")

def run_network_discovery(target_subnet, output_dir="dxso_reports/1_Network_Discovery"):
    os.makedirs(output_dir, exist_ok=True)
    print(f"\n=== PHASE 1: Network Asset Discovery ({target_subnet}) ===")
    
    all_discovered_ips = set()
    safe_subnet_name = target_subnet.replace("/", "_")
    
    # Tool 0: Subnet-specific folder logic
    subnet_out_dir = os.path.join(output_dir, safe_subnet_name)
    os.makedirs(subnet_out_dir, exist_ok=True)
    
    # 1. Nmap basic ping sweep
    nmap_out_file = f"{subnet_out_dir}/nmap_live_hosts.txt"
    cmd_nmap = f"nmap -sn {target_subnet} -oN {nmap_out_file}"
    nmap_out = run_command(cmd_nmap)
    nmap_ips = extract_ips(nmap_out)
    print(f"[*] Nmap discovered {len(nmap_ips)} IPs from {target_subnet}.\n")
    all_discovered_ips.update(nmap_ips)

    # 2. arp-scan
    arp_out_file = f"{subnet_out_dir}/arpscan_hosts.txt"
    cmd_arp_scan = f"arp-scan {target_subnet}"
    arp_scan_out = run_command(cmd_arp_scan)
    with open(arp_out_file, "w") as f:
        f.write(arp_scan_out)
    arp_ips = extract_ips(arp_scan_out)
    print(f"[*] arp-scan discovered {len(arp_ips)} IPs from {target_subnet}.\n")
    all_discovered_ips.update(arp_ips)

    # 3. netdiscover
    netdiscover_out_file = f"{subnet_out_dir}/netdiscover_hosts.txt"
    cmd_netdiscover = f"netdiscover -r {target_subnet} -P -N"
    netdiscover_out = run_command(cmd_netdiscover)
    with open(netdiscover_out_file, "w") as f:
        f.write(netdiscover_out)
    netdiscover_ips = extract_ips(netdiscover_out)
    print(f"[*] netdiscover discovered {len(netdiscover_ips)} IPs from {target_subnet}.\n")
    all_discovered_ips.update(netdiscover_ips)

    # 4. arping sweep
    arping_out_file = f"{subnet_out_dir}/arping_hosts.txt"
    arping_ips = set()
    print(f"[*] Executing arping sweep on {target_subnet}... this might take a moment.")
    try:
        network = ipaddress.ip_network(target_subnet, strict=False)
        with open(arping_out_file, "w") as f:
            for ip in network.hosts():
                cmd_arping = f"arping -c 1 -w 1 {ip} 2>/dev/null"
                res = subprocess.run(cmd_arping, shell=True, capture_output=True, text=True)
                f.write(res.stdout)
                
                output_lower = res.stdout.lower()
                if "reply" in output_lower or "bytes from" in output_lower:
                    arping_ips.add(str(ip))
    except Exception as e:
        print(f"[!] Could not run arping comprehensively: {e}")
        
    print(f"[*] arping discovered {len(arping_ips)} IPs from {target_subnet}.\n")
    all_discovered_ips.update(arping_ips)

    # Collation
    master_list_file = f"{subnet_out_dir}/collated_live_hosts.txt"
    try:
        network = ipaddress.ip_network(target_subnet, strict=False)
        all_discovered_ips.discard(str(network.network_address))
        all_discovered_ips.discard(str(network.broadcast_address))
    except Exception:
        pass 

    sorted_ips = sorted(
        list(all_discovered_ips),
        key=lambda ip: tuple(int(part) for part in ip.split('.')) if ip.count('.') == 3 else ip
    )

    with open(master_list_file, "w") as f:
        f.write(f"=== COLLATED LIVE HOSTS ===\n")
        f.write(f"Target Subnet: {target_subnet}\n")
        f.write(f"Total Unique IPs Discovered: {len(sorted_ips)}\n")
        f.write(f"Tools Used: nmap, arp-scan, netdiscover, arping\n")
        f.write("=" * 27 + "\n\n")
        for ip in sorted_ips:
            f.write(f"{ip}\n")

    print(f"[+] Collation complete for {target_subnet}! Found {len(sorted_ips)} unique IPs.")
    print(f"[+] Master list saved to {master_list_file}")

    # Capture routing details for downstream consumption (Phase 3)
    config_file = f"{output_dir}/network_config.txt"
    with open(config_file, "w") as f:
        # Attempt to grab the default gateway natively
        try:
            route_out = subprocess.check_output("ip route | grep default || netstat -rn | grep default || route -n get default | grep gateway", shell=True, text=True, stderr=subprocess.DEVNULL)
            gateway = route_out.split()[2] if "ip route" in subprocess.check_output("which ip", shell=True, text=True, stderr=subprocess.DEVNULL) else route_out.split()[1] # Naive extraction
            f.write(f"GATEWAY={gateway}\n")
        except Exception:
            f.write("GATEWAY=unknown\n")
            
        # Attempt to snag active interfaces natively
        try:
            iface_out = subprocess.check_output("ip route | grep default || netstat -rn | grep default", shell=True, text=True, stderr=subprocess.DEVNULL)
            iface = iface_out.split()[4] if "dev" in iface_out else iface_out.split()[3] # Naive
            f.write(f"INTERFACE={iface}\n")
        except Exception:
            f.write("INTERFACE=wlan0\n")

    print("\n[*] To perform a detailed phase 1.5 scan on discovered hosts, run:")
    print(f"    nmap -A -T4 -O -iL {master_list_file} -oN {subnet_out_dir}/detailed_scan.txt")
    print("-" * 50)
    
    # Generate the requested Phase 1 Excel Port Matrix
    generate_port_matrix(master_list_file, subnet_out_dir)
    
    return sorted_ips

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Phase 1: Network Asset Discovery")
    parser.add_argument("-t", "--target", help="Target subnet (e.g., 192.168.1.0/24). If omitted, auto-discovery is used.")
    args = parser.parse_args()
    
    targets = []
    
    if args.target:
        targets.append(args.target)
    else:
        # Auto-discover subnets
        discovered = discover_local_subnets()
        if not discovered:
            print("[!] Could not automatically discover any subnets. Please provide one manually with -t.")
            exit(1)
        print(f"[+] Automatically discovered {len(discovered)} local subnet(s): {', '.join(discovered)}")
        targets.extend(discovered)
        
    for index, subnet in enumerate(targets):
        run_network_discovery(subnet)
        
    if len(targets) > 1:
        print(f"\n[+] Total scan sequence completed for {len(targets)} subnets.")
